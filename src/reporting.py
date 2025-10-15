"""
Reporting module for customer support analytics.
Handles automated report generation and multi-format export capabilities.
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Optional, Tuple, Any
import logging
from datetime import datetime, timedelta
import io
import base64
import warnings
warnings.filterwarnings('ignore')

# Import reporting libraries
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logging.warning("ReportLab not available. PDF export will be limited.")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("OpenPyXL not available. Excel export will be limited.")

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ReportGenerator:
    """Handles automated report generation and export functionality."""
    
    def __init__(self):
        """Initialize the report generator."""
        self.report_templates = {
            'executive_summary': self._generate_executive_summary,
            'team_performance': self._generate_team_performance_report,
            'sentiment_analysis': self._generate_sentiment_report,
            'response_time_analysis': self._generate_response_time_report,
            'anomaly_report': self._generate_anomaly_report,
            'comprehensive': self._generate_comprehensive_report
        }
        
        self.export_formats = ['pdf', 'excel', 'csv', 'html']
        
        logger.info("Report generator initialized")
    
    def generate_report(self, data: pd.DataFrame, report_type: str = 'comprehensive', 
                       format: str = 'pdf', **kwargs) -> Dict:
        """
        Generate a report based on the specified type and format.
        
        Args:
            data: DataFrame with customer support data
            report_type: Type of report to generate
            format: Export format (pdf, excel, csv, html)
            **kwargs: Additional parameters for report generation
            
        Returns:
            Dict: Report generation results
        """
        try:
            if report_type not in self.report_templates:
                return {'error': f'Unknown report type: {report_type}'}
            
            if format not in self.export_formats:
                return {'error': f'Unsupported format: {format}'}
            
            # Generate report content
            report_content = self.report_templates[report_type](data, **kwargs)
            
            # Export in specified format
            if format == 'pdf':
                export_result = self._export_to_pdf(report_content, report_type)
            elif format == 'excel':
                export_result = self._export_to_excel(report_content, report_type)
            elif format == 'csv':
                export_result = self._export_to_csv(report_content, report_type)
            elif format == 'html':
                export_result = self._export_to_html(report_content, report_type)
            
            return {
                'success': True,
                'report_type': report_type,
                'format': format,
                'content': report_content,
                'export': export_result,
                'generated_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Error generating report: {e}")
            return {'error': str(e)}
    
    def generate_executive_summary(self, data: pd.DataFrame) -> str:
        """Generate executive summary report."""
        try:
            summary = self._generate_executive_summary(data)
            return summary['content']
        except Exception as e:
            logger.error(f"Error generating executive summary: {e}")
            return f"Error generating executive summary: {e}"
    
    def generate_team_report(self, team_data: pd.DataFrame, team_name: str = None) -> str:
        """Generate team-specific report."""
        try:
            report = self._generate_team_performance_report(team_data, team_name)
            return report['content']
        except Exception as e:
            logger.error(f"Error generating team report: {e}")
            return f"Error generating team report: {e}"
    
    def generate_trend_report(self, trend_data: pd.DataFrame) -> str:
        """Generate trend analysis report."""
        try:
            report = self._generate_response_time_report(trend_data)
            return report['content']
        except Exception as e:
            logger.error(f"Error generating trend report: {e}")
            return f"Error generating trend report: {e}"
    
    def export_to_pdf(self, report_data: Dict) -> bytes:
        """Export report to PDF format."""
        try:
            if not REPORTLAB_AVAILABLE:
                return b"PDF export not available - ReportLab not installed"
            
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            styles = getSampleStyleSheet()
            
            # Create custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=TA_CENTER,
                textColor=colors.darkblue
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=12,
                textColor=colors.darkblue
            )
            
            # Build PDF content
            story = []
            
            # Title
            story.append(Paragraph("Customer Support Analytics Report", title_style))
            story.append(Spacer(1, 12))
            
            # Report metadata
            story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Add report sections
            for section, content in report_data.items():
                if section == 'metadata':
                    continue
                
                story.append(Paragraph(section.replace('_', ' ').title(), heading_style))
                
                if isinstance(content, str):
                    story.append(Paragraph(content, styles['Normal']))
                elif isinstance(content, list):
                    for item in content:
                        story.append(Paragraph(f"• {item}", styles['Normal']))
                elif isinstance(content, dict):
                    # Create table for dictionary data
                    table_data = [['Metric', 'Value']]
                    for key, value in content.items():
                        table_data.append([str(key), str(value)])
                    
                    table = Table(table_data)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 14),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    story.append(table)
                
                story.append(Spacer(1, 20))
            
            # Build PDF
            doc.build(story)
            buffer.seek(0)
            
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error exporting to PDF: {e}")
            return b"Error generating PDF"
    
    def export_to_excel(self, report_data: Dict, filename: str = None) -> bytes:
        """Export report to Excel format."""
        try:
            if not OPENPYXL_AVAILABLE:
                return b"Excel export not available - OpenPyXL not installed"
            
            buffer = io.BytesIO()
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Create summary sheet
            summary_sheet = workbook.create_sheet("Summary")
            
            # Add report metadata
            summary_sheet['A1'] = "Customer Support Analytics Report"
            summary_sheet['A1'].font = Font(size=16, bold=True)
            summary_sheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            row = 4
            
            # Add report sections
            for section, content in report_data.items():
                if section == 'metadata':
                    continue
                
                summary_sheet[f'A{row}'] = section.replace('_', ' ').title()
                summary_sheet[f'A{row}'].font = Font(size=14, bold=True)
                row += 1
                
                if isinstance(content, dict):
                    # Add dictionary as table
                    for key, value in content.items():
                        summary_sheet[f'A{row}'] = str(key)
                        summary_sheet[f'B{row}'] = str(value)
                        row += 1
                elif isinstance(content, list):
                    # Add list items
                    for item in content:
                        summary_sheet[f'A{row}'] = f"• {item}"
                        row += 1
                else:
                    # Add string content
                    summary_sheet[f'A{row}'] = str(content)
                    row += 1
                
                row += 1  # Add spacing
            
            # Auto-adjust column widths
            for column in summary_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                summary_sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save to buffer
            workbook.save(buffer)
            buffer.seek(0)
            
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            return b"Error generating Excel file"
    
    def export_to_csv(self, report_data: Dict) -> str:
        """Export report to CSV format."""
        try:
            csv_content = []
            csv_content.append("Customer Support Analytics Report")
            csv_content.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            csv_content.append("")
            
            for section, content in report_data.items():
                if section == 'metadata':
                    continue
                
                csv_content.append(f"=== {section.replace('_', ' ').title()} ===")
                
                if isinstance(content, dict):
                    for key, value in content.items():
                        csv_content.append(f"{key},{value}")
                elif isinstance(content, list):
                    for item in content:
                        csv_content.append(f"• {item}")
                else:
                    csv_content.append(str(content))
                
                csv_content.append("")
            
            return "\n".join(csv_content)
            
        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            return f"Error generating CSV: {e}"
    
    def export_to_html(self, report_data: Dict) -> str:
        """Export report to HTML format."""
        try:
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Customer Support Analytics Report</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 40px; }}
                    h1 {{ color: #2c3e50; text-align: center; }}
                    h2 {{ color: #34495e; border-bottom: 2px solid #3498db; }}
                    h3 {{ color: #7f8c8d; }}
                    .metadata {{ color: #7f8c8d; text-align: center; margin-bottom: 30px; }}
                    .section {{ margin-bottom: 30px; }}
                    table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
                    th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
                    th {{ background-color: #3498db; color: white; }}
                    tr:nth-child(even) {{ background-color: #f2f2f2; }}
                    ul {{ margin: 10px 0; }}
                    li {{ margin: 5px 0; }}
                </style>
            </head>
            <body>
                <h1>Customer Support Analytics Report</h1>
                <div class="metadata">
                    Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
                </div>
            """
            
            for section, content in report_data.items():
                if section == 'metadata':
                    continue
                
                html_content += f'<div class="section">'
                html_content += f'<h2>{section.replace("_", " ").title()}</h2>'
                
                if isinstance(content, dict):
                    html_content += '<table><tr><th>Metric</th><th>Value</th></tr>'
                    for key, value in content.items():
                        html_content += f'<tr><td>{key}</td><td>{value}</td></tr>'
                    html_content += '</table>'
                elif isinstance(content, list):
                    html_content += '<ul>'
                    for item in content:
                        html_content += f'<li>{item}</li>'
                    html_content += '</ul>'
                else:
                    html_content += f'<p>{content}</p>'
                
                html_content += '</div>'
            
            html_content += """
            </body>
            </html>
            """
            
            return html_content
            
        except Exception as e:
            logger.error(f"Error exporting to HTML: {e}")
            return f"<html><body><h1>Error generating HTML report: {e}</h1></body></html>"
    
    def _generate_executive_summary(self, data: pd.DataFrame) -> Dict:
        """Generate executive summary report content."""
        try:
            summary = {
                'metadata': {
                    'report_type': 'executive_summary',
                    'generated_at': datetime.now().isoformat(),
                    'data_points': len(data)
                },
                'overview': {},
                'key_metrics': {},
                'insights': [],
                'recommendations': []
            }
            
            # Calculate key metrics
            if 'response_time_minutes' in data.columns:
                rt_data = data['response_time_minutes'].dropna()
                summary['key_metrics']['response_time'] = {
                    'median': f"{rt_data.median():.1f} minutes",
                    'average': f"{rt_data.mean():.1f} minutes",
                    'p90': f"{rt_data.quantile(0.9):.1f} minutes",
                    'sla_compliance': f"{((rt_data <= 60).mean() * 100):.1f}%"
                }
            
            if 'combined_score' in data.columns:
                sentiment_data = data['combined_score'].dropna()
                summary['key_metrics']['sentiment'] = {
                    'average_score': f"{sentiment_data.mean():.3f}",
                    'positive_rate': f"{((sentiment_data > 0.05).mean() * 100):.1f}%",
                    'negative_rate': f"{((sentiment_data < -0.05).mean() * 100):.1f}%"
                }
            
            if 'team' in data.columns:
                team_counts = data['team'].value_counts()
                summary['key_metrics']['teams'] = {
                    'total_teams': len(team_counts),
                    'largest_team': team_counts.index[0],
                    'team_distribution': team_counts.to_dict()
                }
            
            # Generate insights
            insights = []
            if 'response_time_minutes' in data.columns:
                rt_median = data['response_time_minutes'].median()
                if rt_median < 30:
                    insights.append("Excellent response time performance")
                elif rt_median < 60:
                    insights.append("Good response time performance")
                else:
                    insights.append("Response times need improvement")
            
            if 'combined_score' in data.columns:
                avg_sentiment = data['combined_score'].mean()
                if avg_sentiment > 0.1:
                    insights.append("Positive customer sentiment trend")
                elif avg_sentiment < -0.1:
                    insights.append("Negative customer sentiment trend")
                else:
                    insights.append("Neutral customer sentiment")
            
            summary['insights'] = insights
            
            # Generate recommendations
            recommendations = []
            if 'response_time_minutes' in data.columns:
                rt_median = data['response_time_minutes'].median()
                if rt_median > 60:
                    recommendations.append("Implement response time optimization strategies")
            
            if 'team' in data.columns:
                recommendations.append("Review team performance and provide targeted training")
            
            recommendations.append("Continue monitoring key performance indicators")
            summary['recommendations'] = recommendations
            
            return summary
            
        except Exception as e:
            logger.error(f"Error generating executive summary: {e}")
            return {'error': str(e)}
    
    def _generate_team_performance_report(self, data: pd.DataFrame, team_name: str = None) -> Dict:
        """Generate team performance report content."""
        try:
            report = {
                'metadata': {
                    'report_type': 'team_performance',
                    'team_name': team_name or 'All Teams',
                    'generated_at': datetime.now().isoformat()
                },
                'team_metrics': {},
                'performance_analysis': {},
                'recommendations': []
            }
            
            if team_name:
                team_data = data[data['team'] == team_name]
            else:
                team_data = data
            
            # Calculate team metrics
            if 'response_time_minutes' in team_data.columns:
                rt_data = team_data['response_time_minutes'].dropna()
                report['team_metrics']['response_time'] = {
                    'median': f"{rt_data.median():.1f} minutes",
                    'average': f"{rt_data.mean():.1f} minutes",
                    'std_deviation': f"{rt_data.std():.1f} minutes",
                    'min': f"{rt_data.min():.1f} minutes",
                    'max': f"{rt_data.max():.1f} minutes"
                }
            
            if 'combined_score' in team_data.columns:
                sentiment_data = team_data['combined_score'].dropna()
                report['team_metrics']['sentiment'] = {
                    'average_score': f"{sentiment_data.mean():.3f}",
                    'positive_rate': f"{((sentiment_data > 0.05).mean() * 100):.1f}%",
                    'negative_rate': f"{((sentiment_data < -0.05).mean() * 100):.1f}%",
                    'neutral_rate': f"{((sentiment_data >= -0.05) & (sentiment_data <= 0.05)).mean() * 100:.1f}%"
                }
            
            # Performance analysis
            if 'team' in data.columns and team_name:
                # Compare with other teams
                other_teams = data[data['team'] != team_name]
                if not other_teams.empty:
                    comparison = self._compare_team_performance(team_data, other_teams)
                    report['performance_analysis']['comparison'] = comparison
            
            # Generate recommendations
            recommendations = []
            if 'response_time_minutes' in team_data.columns:
                rt_median = team_data['response_time_minutes'].median()
                if rt_median > 60:
                    recommendations.append("Focus on reducing response times")
                elif rt_median < 30:
                    recommendations.append("Maintain excellent response time performance")
            
            if 'combined_score' in team_data.columns:
                avg_sentiment = team_data['combined_score'].mean()
                if avg_sentiment < 0:
                    recommendations.append("Improve customer communication and service quality")
            
            recommendations.append("Continue monitoring team performance metrics")
            report['recommendations'] = recommendations
            
            return report
            
        except Exception as e:
            logger.error(f"Error generating team performance report: {e}")
            return {'error': str(e)}
    
    def _generate_sentiment_report(self, data: pd.DataFrame) -> Dict:
        """Generate sentiment analysis report content."""
        try:
            report = {
                'metadata': {
                    'report_type': 'sentiment_analysis',
                    'generated_at': datetime.now().isoformat()
                },
                'sentiment_overview': {},
                'trend_analysis': {},
                'team_sentiment': {},
                'recommendations': []
            }
            
            if 'combined_score' not in data.columns:
                return {'error': 'No sentiment data available'}
            
            sentiment_data = data['combined_score'].dropna()
            
            # Sentiment overview
            report['sentiment_overview'] = {
                'total_messages': len(sentiment_data),
                'average_score': f"{sentiment_data.mean():.3f}",
                'score_std': f"{sentiment_data.std():.3f}",
                'positive_rate': f"{((sentiment_data > 0.05).mean() * 100):.1f}%",
                'negative_rate': f"{((sentiment_data < -0.05).mean() * 100):.1f}%",
                'neutral_rate': f"{((sentiment_data >= -0.05) & (sentiment_data <= 0.05)).mean() * 100:.1f}%"
            }
            
            # Trend analysis
            if 'created_at' in data.columns:
                trend_analysis = self._analyze_sentiment_trends(data)
                report['trend_analysis'] = trend_analysis
            
            # Team sentiment analysis
            if 'team' in data.columns:
                team_sentiment = {}
                for team in data['team'].unique():
                    team_data = data[data['team'] == team]
                    team_sentiment_data = team_data['combined_score'].dropna()
                    
                    if len(team_sentiment_data) > 0:
                        team_sentiment[team] = {
                            'average_score': f"{team_sentiment_data.mean():.3f}",
                            'positive_rate': f"{((team_sentiment_data > 0.05).mean() * 100):.1f}%",
                            'message_count': len(team_sentiment_data)
                        }
                
                report['team_sentiment'] = team_sentiment
            
            # Generate recommendations
            recommendations = []
            avg_sentiment = sentiment_data.mean()
            
            if avg_sentiment > 0.1:
                recommendations.append("Maintain positive customer sentiment through consistent service quality")
            elif avg_sentiment < -0.1:
                recommendations.append("Implement customer satisfaction improvement initiatives")
            else:
                recommendations.append("Focus on converting neutral sentiment to positive")
            
            recommendations.append("Monitor sentiment trends regularly")
            recommendations.append("Provide team-specific sentiment training")
            
            report['recommendations'] = recommendations
            
            return report
            
        except Exception as e:
            logger.error(f"Error generating sentiment report: {e}")
            return {'error': str(e)}
    
    def _generate_response_time_report(self, data: pd.DataFrame) -> Dict:
        """Generate response time analysis report content."""
        try:
            report = {
                'metadata': {
                    'report_type': 'response_time_analysis',
                    'generated_at': datetime.now().isoformat()
                },
                'response_time_metrics': {},
                'sla_analysis': {},
                'team_performance': {},
                'recommendations': []
            }
            
            if 'response_time_minutes' not in data.columns:
                return {'error': 'No response time data available'}
            
            rt_data = data['response_time_minutes'].dropna()
            
            # Response time metrics
            report['response_time_metrics'] = {
                'total_tickets': len(rt_data),
                'median': f"{rt_data.median():.1f} minutes",
                'average': f"{rt_data.mean():.1f} minutes",
                'p90': f"{rt_data.quantile(0.9):.1f} minutes",
                'p95': f"{rt_data.quantile(0.95):.1f} minutes",
                'std_deviation': f"{rt_data.std():.1f} minutes",
                'min': f"{rt_data.min():.1f} minutes",
                'max': f"{rt_data.max():.1f} minutes"
            }
            
            # SLA analysis
            sla_breach_rate = (rt_data > 60).mean() * 100
            report['sla_analysis'] = {
                'sla_threshold': '60 minutes',
                'breach_rate': f"{sla_breach_rate:.1f}%",
                'compliance_rate': f"{100 - sla_breach_rate:.1f}%",
                'breach_count': int((rt_data > 60).sum()),
                'status': 'Good' if sla_breach_rate < 10 else 'Needs Improvement'
            }
            
            # Team performance
            if 'team' in data.columns:
                team_performance = {}
                for team in data['team'].unique():
                    team_data = data[data['team'] == team]
                    team_rt = team_data['response_time_minutes'].dropna()
                    
                    if len(team_rt) > 0:
                        team_performance[team] = {
                            'median_rt': f"{team_rt.median():.1f} minutes",
                            'sla_compliance': f"{((team_rt <= 60).mean() * 100):.1f}%",
                            'ticket_count': len(team_rt)
                        }
                
                report['team_performance'] = team_performance
            
            # Generate recommendations
            recommendations = []
            
            if sla_breach_rate > 20:
                recommendations.append("Urgent: Implement response time optimization strategies")
            elif sla_breach_rate > 10:
                recommendations.append("Focus on reducing response times to improve SLA compliance")
            else:
                recommendations.append("Maintain current response time performance")
            
            recommendations.append("Review team-specific response time patterns")
            recommendations.append("Implement response time monitoring and alerting")
            
            report['recommendations'] = recommendations
            
            return report
            
        except Exception as e:
            logger.error(f"Error generating response time report: {e}")
            return {'error': str(e)}
    
    def _generate_anomaly_report(self, data: pd.DataFrame) -> Dict:
        """Generate anomaly detection report content."""
        try:
            # Import anomaly detector
            try:
                from anomaly_detection import AnomalyDetector
                detector = AnomalyDetector()
                anomalies = detector.detect_anomalies(data)
            except ImportError:
                return {'error': 'Anomaly detection module not available'}
            
            report = {
                'metadata': {
                    'report_type': 'anomaly_detection',
                    'generated_at': datetime.now().isoformat()
                },
                'anomaly_summary': {},
                'detailed_analysis': {},
                'recommendations': []
            }
            
            # Anomaly summary
            summary = anomalies.get('summary', {})
            report['anomaly_summary'] = {
                'total_anomalies': summary.get('total_anomalies', 0),
                'overall_severity': summary.get('overall_severity', 'unknown'),
                'anomaly_types': summary.get('anomaly_types', [])
            }
            
            # Detailed analysis
            for anomaly_type, anomaly_data in anomalies.items():
                if anomaly_type != 'summary' and isinstance(anomaly_data, dict):
                    report['detailed_analysis'][anomaly_type] = {
                        'count': anomaly_data.get('count', 0),
                        'percentage': anomaly_data.get('percentage', 0),
                        'severity': anomaly_data.get('severity', 'unknown')
                    }
            
            # Recommendations
            recommendations = summary.get('recommendations', [])
            if not recommendations:
                recommendations = ["No significant anomalies detected - maintain current monitoring"]
            
            report['recommendations'] = recommendations
            
            return report
            
        except Exception as e:
            logger.error(f"Error generating anomaly report: {e}")
            return {'error': str(e)}
    
    def _generate_comprehensive_report(self, data: pd.DataFrame) -> Dict:
        """Generate comprehensive report combining all analyses."""
        try:
            report = {
                'metadata': {
                    'report_type': 'comprehensive',
                    'generated_at': datetime.now().isoformat(),
                    'data_points': len(data)
                },
                'executive_summary': {},
                'response_time_analysis': {},
                'sentiment_analysis': {},
                'team_performance': {},
                'anomaly_detection': {},
                'recommendations': []
            }
            
            # Generate all sub-reports
            exec_summary = self._generate_executive_summary(data)
            report['executive_summary'] = exec_summary.get('key_metrics', {})
            
            rt_report = self._generate_response_time_report(data)
            report['response_time_analysis'] = rt_report.get('response_time_metrics', {})
            
            sentiment_report = self._generate_sentiment_report(data)
            report['sentiment_analysis'] = sentiment_report.get('sentiment_overview', {})
            
            team_report = self._generate_team_performance_report(data)
            report['team_performance'] = team_report.get('team_metrics', {})
            
            # Anomaly detection
            try:
                from anomaly_detection import AnomalyDetector
                detector = AnomalyDetector()
                anomalies = detector.detect_anomalies(data)
                report['anomaly_detection'] = anomalies.get('summary', {})
            except ImportError:
                report['anomaly_detection'] = {'error': 'Anomaly detection not available'}
            
            # Combine recommendations
            all_recommendations = []
            all_recommendations.extend(exec_summary.get('recommendations', []))
            all_recommendations.extend(rt_report.get('recommendations', []))
            all_recommendations.extend(sentiment_report.get('recommendations', []))
            all_recommendations.extend(team_report.get('recommendations', []))
            
            # Remove duplicates and limit
            unique_recommendations = list(dict.fromkeys(all_recommendations))
            report['recommendations'] = unique_recommendations[:10]
            
            return report
            
        except Exception as e:
            logger.error(f"Error generating comprehensive report: {e}")
            return {'error': str(e)}
    
    def _compare_team_performance(self, team_data: pd.DataFrame, other_teams: pd.DataFrame) -> Dict:
        """Compare team performance with other teams."""
        try:
            comparison = {}
            
            if 'response_time_minutes' in team_data.columns and 'response_time_minutes' in other_teams.columns:
                team_rt = team_data['response_time_minutes'].dropna()
                other_rt = other_teams['response_time_minutes'].dropna()
                
                if len(team_rt) > 0 and len(other_rt) > 0:
                    comparison['response_time'] = {
                        'team_median': f"{team_rt.median():.1f} minutes",
                        'others_median': f"{other_rt.median():.1f} minutes",
                        'performance_vs_others': 'Better' if team_rt.median() < other_rt.median() else 'Worse'
                    }
            
            if 'combined_score' in team_data.columns and 'combined_score' in other_teams.columns:
                team_sentiment = team_data['combined_score'].dropna()
                other_sentiment = other_teams['combined_score'].dropna()
                
                if len(team_sentiment) > 0 and len(other_sentiment) > 0:
                    comparison['sentiment'] = {
                        'team_average': f"{team_sentiment.mean():.3f}",
                        'others_average': f"{other_sentiment.mean():.3f}",
                        'performance_vs_others': 'Better' if team_sentiment.mean() > other_sentiment.mean() else 'Worse'
                    }
            
            return comparison
            
        except Exception as e:
            logger.error(f"Error comparing team performance: {e}")
            return {}
    
    def _analyze_sentiment_trends(self, data: pd.DataFrame) -> Dict:
        """Analyze sentiment trends over time."""
        try:
            if 'created_at' not in data.columns or 'combined_score' not in data.columns:
                return {'error': 'Insufficient data for trend analysis'}
            
            # Group by date and calculate daily sentiment
            data['date'] = pd.to_datetime(data['created_at']).dt.date
            daily_sentiment = data.groupby('date')['combined_score'].agg(['mean', 'count']).reset_index()
            
            if len(daily_sentiment) < 2:
                return {'error': 'Insufficient data points for trend analysis'}
            
            # Calculate trend
            sentiment_values = daily_sentiment['mean'].values
            trend_slope = np.polyfit(range(len(sentiment_values)), sentiment_values, 1)[0]
            
            trend_direction = 'improving' if trend_slope > 0.001 else 'declining' if trend_slope < -0.001 else 'stable'
            
            return {
                'trend_direction': trend_direction,
                'trend_slope': f"{trend_slope:.6f}",
                'data_points': len(daily_sentiment),
                'average_daily_sentiment': f"{daily_sentiment['mean'].mean():.3f}",
                'sentiment_volatility': f"{daily_sentiment['mean'].std():.3f}"
            }
            
        except Exception as e:
            logger.error(f"Error analyzing sentiment trends: {e}")
            return {'error': str(e)}
    
    def _export_to_pdf(self, report_content: Dict, report_type: str) -> Dict:
        """Export report content to PDF."""
        try:
            pdf_bytes = self.export_to_pdf(report_content)
            
            return {
                'success': True,
                'format': 'pdf',
                'size_bytes': len(pdf_bytes),
                'download_url': f"data:application/pdf;base64,{base64.b64encode(pdf_bytes).decode()}"
            }
            
        except Exception as e:
            logger.error(f"Error exporting to PDF: {e}")
            return {'success': False, 'error': str(e)}
    
    def _export_to_excel(self, report_content: Dict, report_type: str) -> Dict:
        """Export report content to Excel."""
        try:
            excel_bytes = self.export_to_excel(report_content)
            
            return {
                'success': True,
                'format': 'excel',
                'size_bytes': len(excel_bytes),
                'download_url': f"data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{base64.b64encode(excel_bytes).decode()}"
            }
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            return {'success': False, 'error': str(e)}
    
    def _export_to_csv(self, report_content: Dict, report_type: str) -> Dict:
        """Export report content to CSV."""
        try:
            csv_content = self.export_to_csv(report_content)
            
            return {
                'success': True,
                'format': 'csv',
                'content': csv_content,
                'size_bytes': len(csv_content.encode('utf-8'))
            }
            
        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            return {'success': False, 'error': str(e)}
    
    def _export_to_html(self, report_content: Dict, report_type: str) -> Dict:
        """Export report content to HTML."""
        try:
            html_content = self.export_to_html(report_content)
            
            return {
                'success': True,
                'format': 'html',
                'content': html_content,
                'size_bytes': len(html_content.encode('utf-8'))
            }
            
        except Exception as e:
            logger.error(f"Error exporting to HTML: {e}")
            return {'success': False, 'error': str(e)}
